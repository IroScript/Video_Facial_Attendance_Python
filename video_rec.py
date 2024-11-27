import os
import datetime
import tkinter as tk
import cv2
from PIL import Image, ImageTk
import util
import openpyxl
import ntplib
import face_recognition
import time

class App:
    def __init__(self):
        self.main_window = tk.Tk()
        self.center_window(self.main_window, 1200, 520)

        self.login_button_main_window = util.get_button(self.main_window, 'LOGIN', 'green', self.login)
        self.login_button_main_window.place(x=750, y=300)

        self.register_new_user_button_main_window = util.get_button(self.main_window, 'REGISTER NEW USER', 'gray', 
                                                                    self.register_new_user, fg='black')
        self.register_new_user_button_main_window.place(x=750, y=400)

        self.webcam_label = util.get_img_label(self.main_window)
        self.webcam_label.place(x=10, y=0, width=700, height=500)

        self.db_dir = './db'
        if not os.path.exists(self.db_dir):
            os.mkdir(self.db_dir)

        self.initialize_excel()

        self.cap = cv2.VideoCapture(0)
        self.video_frames = []
        self.is_capturing = False
        self.capture_start_time = None
        self.countdown_label = tk.Label(self.main_window, text="", font=('Arial', 20), bg='white')
        self.countdown_label.place(x=750, y=50)

        self.add_webcam(self.webcam_label)

    def center_window(self, window, width, height):
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        window.geometry(f"{width}x{height}+{x}+{y}")

    def add_webcam(self, label):
        self._label = label
        self.process_webcam()

    def process_webcam(self):
        ret, frame = self.cap.read()
        if ret:
            if self.is_capturing:
                self.video_frames.append(frame)
                current_time = time.time()
                time_elapsed = current_time - self.capture_start_time
                
                if self.capture_mode == 'register' and time_elapsed >= 30:
                    self.stop_capture()
                    self.complete_registration()
                elif self.capture_mode == 'login':
                    face_locations = face_recognition.face_locations(frame)
                    if face_locations:
                        self.stop_capture()
                        self.process_login()

            img_ = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            self.most_recent_capture_pil = Image.fromarray(img_)
            imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
            self._label.imgtk = imgtk
            self._label.configure(image=imgtk)

        self._label.after(20, self.process_webcam)

    def start_capture(self, mode):
        self.is_capturing = True
        self.capture_start_time = time.time()
        self.video_frames = []
        self.capture_mode = mode
        if mode == 'register':
            self.update_countdown(30)
        else:
            self.countdown_label.config(text="")

    def update_countdown(self, seconds):
        if self.capture_mode == 'register':
            self.countdown_label.config(text=f"Please stay here! Capture in: {seconds} seconds")
            if seconds > 0:
                self.main_window.after(1000, self.update_countdown, seconds - 1)
            else:
                self.countdown_label.config(text="Capture complete!")

    def stop_capture(self):
        self.is_capturing = False

    def verify_time(self):
        try:
            ntp_client = ntplib.NTPClient()
            response = ntp_client.request('pool.ntp.org', version=3)
            current_time = datetime.datetime.fromtimestamp(response.tx_time)
            system_time = datetime.datetime.now()
            time_difference = abs((current_time - system_time).total_seconds())
            if time_difference > 60:  # Time difference exceeds 1 minute
                util.msg_box('TIME ERROR', 'TIME IS NOT UPDATED. PLEASE UPDATE TIME.')
                return False
            return True
        except:
            return True  # Assume correct time for offline cases

    def login(self):
        if not self.verify_time():
            return

        self.start_capture('login')

    def process_login(self):
        if not self.video_frames:
            util.msg_box('ERROR', 'No video captured. Please try again.')
            return

        known_faces = []
        known_names = []

        for filename in os.listdir(self.db_dir):
            if filename.endswith('.mp4'):
                name = os.path.splitext(filename)[0]
                video = cv2.VideoCapture(os.path.join(self.db_dir, filename))
                ret, frame = video.read()
                if ret:
                    encodings = face_recognition.face_encodings(frame)
                    if encodings:
                        known_faces.append(encodings[0])
                        known_names.append(name)
                video.release()

        for frame in self.video_frames:
            unknown_encodings = face_recognition.face_encodings(frame)
            if unknown_encodings:
                unknown_encoding = unknown_encodings[0]
                results = face_recognition.compare_faces(known_faces, unknown_encoding)
                if True in results:
                    name = known_names[results.index(True)]
                    now = datetime.datetime.now()
                    login_type = self.determine_login_type(name)
                    self.save_login_video(name, login_type)
                    self.show_popup(f"{name.upper()}, YOUR {login_type} IS {now.strftime('%I:%M %p')}",
                                    "LOGIN SUCCESSFUL")
                    self.update_excel(name, login_type)
                    return

        util.msg_box('ACCESS DENIED', 'UNKNOWN USER. PLEASE REGISTER NEW USER OR TRY AGAIN.')

    def determine_login_type(self, name):
        now = datetime.datetime.now()
        current_date = now.date()
        user_folder = os.path.join(self.db_dir, name, str(now.year), now.strftime("%B"))

        if os.path.exists(user_folder):
            for file in os.listdir(user_folder):
                if current_date.strftime('%d %b %Y') in file and "IN TIME" in file:
                    return "OUT TIME"
        return "IN TIME"

    def show_popup(self, message, title="Information"):
        popup_window = tk.Toplevel(self.main_window)
        popup_window.title(title)
        tk.Label(popup_window, text=message, padx=20, pady=20).pack()
        tk.Button(popup_window, text="OK", command=popup_window.destroy).pack(pady=10)

    def save_login_video(self, name, login_type):
        now = datetime.datetime.now()
        year_folder = os.path.join(self.db_dir, name, str(now.year))
        month_folder = os.path.join(year_folder, now.strftime("%B"))
        os.makedirs(month_folder, exist_ok=True)

        filename = f"{name.upper()}, {now.strftime('%d %b %Y')}, {login_type}, {now.strftime('%I.%M %p')}.mp4"
        filepath = os.path.join(month_folder, filename)

        if login_type == "OUT TIME":
            for existing_video in os.listdir(month_folder):
                if now.strftime('%d %b %Y') in existing_video and "OUT TIME" in existing_video:
                    os.remove(os.path.join(month_folder, existing_video))
                    break

        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(filepath, fourcc, 20.0, (640, 480))
        for frame in self.video_frames:
            out.write(frame)
        out.release()

    def register_new_user(self):
        self.register_new_user_window = tk.Toplevel(self.main_window)
        self.center_window(self.register_new_user_window, 1200, 520)

        self.accept_button_register_new_user_window = util.get_button(self.register_new_user_window, 'ACCEPT', 'green', self.start_registration)
        self.accept_button_register_new_user_window.place(x=750, y=300)

        self.try_again_button_register_new_user_window = util.get_button(self.register_new_user_window, 'TRY AGAIN', 'red', self.try_again_register_new_user)
        self.try_again_button_register_new_user_window.place(x=750, y=400)

        self.capture_label = util.get_img_label(self.register_new_user_window)
        self.capture_label.place(x=10, y=0, width=700, height=500)

        self.entry_text_register_new_user = util.get_entry_text(self.register_new_user_window)
        self.entry_text_register_new_user.place(x=750, y=150)

        self.text_label_register_new_user = util.get_text_label(self.register_new_user_window, 'PLEASE, \nINPUT USERNAME')
        self.text_label_register_new_user.place(x=750, y=70)

    def start_registration(self):
        name = self.entry_text_register_new_user.get("1.0", "end-1c").upper()
        name = name.replace("\r", "").replace("\n", "").strip()

        if not name:
            util.msg_box('ERROR', 'Please enter a username.')
            return
        
        # Close the registration window first 
        # self.register_new_user_window.destroy()

        self.start_capture('register')
        self.update_register_new_user_window()
        self.update_countdown(30)

    def update_register_new_user_window(self):
        if self.is_capturing and self.video_frames:
            frame = self.video_frames[-1]
            img_ = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img_ = Image.fromarray(img_)
            imgtk = ImageTk.PhotoImage(image=img_)
            self.capture_label.imgtk = imgtk
            self.capture_label.configure(image=imgtk)

        if self.is_capturing:
            self.register_new_user_window.after(20, self.update_register_new_user_window)

    def try_again_register_new_user(self):
        self.stop_capture()
        self.register_new_user_window.destroy()

    def complete_registration(self):
        name = self.entry_text_register_new_user.get("1.0", "end-1c").upper()
        name = name.replace("\r", "").replace("\n", "").strip()

        if not self.video_frames:
            util.msg_box('ERROR', 'No video captured. Please try again.')
            return

        face_detected = False
        for frame in self.video_frames:
            face_locations = face_recognition.face_locations(frame)
            if face_locations:
                face_detected = True
                break

        if not face_detected:
            util.msg_box('ERROR', 'No face detected in the video. Please try again.')
            return

        user_video_path = os.path.join(self.db_dir, f'{name}.mp4')
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(user_video_path, fourcc, 20.0, (640, 480))
        for frame in self.video_frames:
            out.write(frame)
        out.release()

        self.show_popup('User was registered successfully!')
        self.register_new_user_window.destroy()

    def initialize_excel(self):
        attendance_report_dir = os.path.join(self.db_dir, "ATTENDANCE REPORT")
        if not os.path.exists(attendance_report_dir):
            os.mkdir(attendance_report_dir)

        current_year = datetime.datetime.now().year
        year_folder = os.path.join(attendance_report_dir, str(current_year))
        if not os.path.exists(year_folder):
            os.mkdir(year_folder)

        current_month_name = datetime.datetime.now().strftime("%B %Y")
        self.excel_file_path = os.path.join(year_folder, f"ATTENDANCE REPORT, {current_month_name}.xlsx")

        if not os.path.exists(self.excel_file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = current_month_name
            ws.cell(row=1, column=1, value="USERNAME")
            for day in range(1, 32):
                ws.cell(row=1, column=(day - 1) * 2 + 2, value=f"{day} IN TIME")
                ws.cell(row=1, column=(day - 1) * 2 + 3, value=f"{day} OUT TIME")
            wb.save(self.excel_file_path)

    def update_excel(self, name, login_type):
        wb = openpyxl.load_workbook(self.excel_file_path)
        current_month = datetime.datetime.now().strftime("%B %Y")
        ws = wb[current_month]

        row = self.find_or_create_user_row(ws, name)
        day = datetime.datetime.now().day
        col = (day - 1) * 2 + (2 if login_type == "IN TIME" else 3)

        if login_type == "IN TIME":
            current_value = ws.cell(row=row, column=col).value
            if not current_value:
                ws.cell(row=row, column=col, value=datetime.datetime.now().strftime("%I:%M %p"))
        else:  # For "OUT TIME", always replace
            ws.cell(row=row, column=col, value=datetime.datetime.now().strftime("%I:%M %p"))

        wb.save(self.excel_file_path)

    def find_or_create_user_row(self, ws, name):
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == name:
                return row
        
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=name)
        return next_row


    def start(self):
        self.main_window.mainloop()

if __name__ == "__main__":
    app = App()
    app.start()

    